from abc import ABC, abstractmethod
import configparser
import io
import sys

import openpyxl
import xlrd
from docx import Document

from bs4 import BeautifulSoup

from zipfile import ZipFile
from zipfile import BadZipfile

from rarfile import RarFile
from rarfile import BadRarFile, NotRarFile, NeedFirstVolume

from src.utils import *


config = configparser.ConfigParser()
config.read('config.ini')
MAX_TABLE_NAME = int(config['tables']['MAX_TABLE_NAME'])
GRAPH_ERRORS = config['database']['GRAPH_ERRORS'].split(',')


class Parser(ABC):
    """базовый класс для парсеров документов различных расширений"""
    tables_info = None
    
    def __init__(self, binary=None, html=None):
        self.binary = io.BytesIO(binary)
        self.html = html
    
    @abstractmethod
    def get_tables_info(self):
        pass


class ParserXLSX(Parser):
    def __init__(self, binary, html):
        super().__init__(binary, html)
        try:
            self.workbook = openpyxl.load_workbook(self.binary)
            self.tables_info = self.get_tables_info()
        except (TypeError, openpyxl.utils.exceptions.InvalidFileException, OSError, BadZipfile):
            message = str(sys.exc_info()[1])
            raise ValueError(f'xlsx-файл не может быть прочитан ({message})')
    
    def get_tables_info(self):
        tables_info = []
        for i, worksheet in enumerate(self.workbook.worksheets):
            table = TableObject(i)
            table.name = self.__get_table_name(worksheet)
            table.n_rows = self.__get_n_rows(worksheet)
            table.n_columns = self.__get_n_columns(worksheet)
            table.unit = find_unit_in_table_name(table.name)
            table.number = find_number_in_table_name(table.name)
            tables_info.append(table)
        return tables_info
    
    def __get_table_name(self, worksheet):
        name = ''
        n_rows = self.__get_n_rows(worksheet)
        n_columns = self.__get_n_columns(worksheet)
        
        # собирать название до тех пор, пока в строке не появится больше одной ячейки с текстом
        for i in range(1, n_rows + 1):
            empty_cells = 0
            cells_text = ''
            for j in range(1, n_columns + 1):
                if worksheet.cell(row=i, column=j).value == None:
                    empty_cells += 1
                else:
                    cells_text += str(worksheet.cell(row=i, column=j).value) + ' '
            if empty_cells >= n_columns - 1:
                name += cells_text
            else:
                break
        
        name = clean_text(name)
        return name

    def __get_n_rows(self, worksheet):
        n_rows = 0
        for row in worksheet.rows:
            if not all([cell.value == None for cell in row]):
                n_rows += 1
        return n_rows
    
    def __get_n_columns(self, worksheet):
        n_columns = 0
        for column in worksheet.columns:
            if not all([cell.value == None for cell in column]):
                n_columns += 1
        return n_columns


class ParserXLS(Parser):
    def __init__(self, binary, html):
        super().__init__(binary, html)
        try:
            self.workbook = xlrd.open_workbook(file_contents=self.binary.read())
            self.tables_info = self.get_tables_info()
        except xlrd.XLRDError:
            message = str(sys.exc_info()[1])
            raise ValueError(f'xls-файл не может быть прочитан ({message})')        
    
    def get_tables_info(self):
        tables_info = []
        for i, worksheet in enumerate(self.workbook.sheets()):
            table = TableObject(i)
            table.name = self.__get_table_name(worksheet)
            table.n_rows = worksheet.nrows
            table.n_columns = worksheet.ncols
            table.unit = find_unit_in_table_name(table.name)
            table.number = find_number_in_table_name(table.name)
            tables_info.append(table)
        return tables_info
    
    def __get_table_name(self, worksheet):
        name = ''
        
        nrows = worksheet.nrows
        ncols = worksheet.ncols
        
        # собирать название до тех пор, пока в строке не появится больше одной ячейки с текстом
        for i in range(nrows):
            empty_cells = 0
            cells_text = ''
            for j in range(ncols):
                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                if worksheet.cell_type(rowx=i, colx=j) == 0:
                    empty_cells += 1
                elif worksheet.cell_type(rowx=i, colx=j) == 1:
                    cells_text += worksheet.cell_value(rowx=i, colx=j) + ' '
            if empty_cells >= ncols - 1:
                name += cells_text
            else:
                break
        
        name = clean_text(name)
        return name


class ParserDOCX(Parser):
    def __init__(self, binary, html):
        super().__init__(binary, html)
        try:
            self.document = Document(self.binary)
            self.tables_info = self.get_tables_info()
        except (ValueError, BadZipfile):
            message = str(sys.exc_info()[1])
            raise ValueError(f'docx-файл не может быть прочитан ({message})')
    
    def get_tables_info(self):
        tables_info = []
        
        tables = self.document.tables
        table_names = self.__get_table_name(self.document)
        
        for i, tbl in enumerate(tables):
            table = TableObject(i)
            table.name = table_names[i]
            table.n_rows = len(tbl.rows)
            table.n_columns = len(tbl.columns)
            table.unit = find_unit_in_table_name(table.name)
            table.number = find_number_in_table_name(table.name)
            tables_info.append(table)
        
        return tables_info
    
    def __get_table_name(self, document):
        table_names = []
        previous_text = []
        
        for block in iterate_paragraphs_and_tables(document):
            if isinstance(block, Paragraph):
                text = clean_text(block.text)
                previous_text.append(text)
            elif isinstance(block, Table):
                table_name = ''
            
                # собираем название таблицы в обратном порядке, пока не наткнемся на пустую
                # строку или строку, начинающуюся с цифры (считаем, что это номер таблицы)
                if any(check_starts_with_number(text) for text in previous_text):
                    for text in reversed(previous_text):
                        if check_starts_with_number(text):
                            table_name = ' '.join((text, table_name))
                            break
                        elif text == '' and table_name != '':
                            break
                        else:
                            table_name = ' '.join((text, table_name))
            
                # или же просто возьмем текст предыдущего абзаца
                elif previous_text:
                    table_name = previous_text[-1]
            
                # если собранное название слишком длинное
                if len(table_name) > MAX_TABLE_NAME:
                    table_name = previous_text[-1]
            
                table_name = clean_text(table_name)
                table_names.append(table_name)
                previous_text = []
    
        return table_names


class ParserHTM(Parser):
    def __init__(self, binary, html):
        super().__init__(binary, html)
        if html:
            self.soup = BeautifulSoup(html, 'lxml')
        elif binary:
            self.soup = BeautifulSoup(binary, 'lxml')
        else:
            raise ValueError('HTML-код не собран')
        self.tables_info = self.get_tables_info()
    
    def get_tables_info(self):
        tables_info = []
        
        tables = self.soup.find_all('table')
        table_names = self.__get_table_name(self.soup)
        
        for i, tbl in enumerate(tables):
            table = TableObject(i)
            table.name = table_names[i]
            table.n_rows = self.__get_table_size(tbl)[0]
            table.n_columns = self.__get_table_size(tbl)[1]
            table.unit = find_unit_in_table_name(table.name)
            table.number = find_number_in_table_name(table.name)
            tables_info.append(table)
        
        return tables_info
    
    def __get_table_name(self, soup):
        table_names = []
        previous_text = []
        
        for elm in soup.select('h2, p:not(table p), table'): # css селектор
            if elm.name != 'table':
                text = clean_text(elm.text)
                previous_text.append(text)
            else:
                table_name = ''
                last_not_empty = ''
                
                for text in reversed(previous_text):
                    if text != '':
                        last_not_empty = text
                        break
                
                # собираем название таблицы, пока не наткнемся на пустую строку
                for text in reversed(previous_text):
                    if text == '' and table_name != '':
                        break
                    else:
                        table_name = ' '.join((text, table_name))
                
                # если собранное название слишком длинное
                if len(table_name) > MAX_TABLE_NAME:
                    table_name = last_not_empty
                
                table_name = clean_text(table_name)
                table_names.append(table_name)
                previous_text = []
        
        return table_names
    
    def __get_table_size(self, table):
        max_rows = len(table.find_all('tr'))
        max_columns = 0
        for row in table.find_all('tr'):
            columns = len(row.find_all('td'))
            if columns > max_columns:
                max_columns = columns
        return (max_rows, max_columns)


class ParserArchive:
    def __init__(self, binary, html=None):
        self.binary = io.BytesIO(binary)
        self.archive_file = self.__load_archive()
        self.tables_info = []
        self.failures = {}
    
    def __load_archive(self):
        try:
            self.archive_file = ZipFile(self.binary, 'r')
        except BadZipfile:
            try:
                self.archive_file = RarFile(self.binary, 'r')
            except (BadRarFile, NotRarFile, io.UnsupportedOperation, NeedFirstVolume):
                message = str(sys.exc_info()[1])
                raise ValueError(f'архив не может быть прочитан ({message})')
        return self.archive_file
    
    def get_tables_info(self):
        for file_name in self.archive_file.namelist():
            file_ext = self.__get_file_ext(file_name)
            if file_ext:
                try:
                    self.__append_tables_info(self.archive_file, file_name, file_ext)
                except TypeError:
                    message = str(sys.exc_info()[1])
                    self.failures[file_name] = message
                    continue
        return self.tables_info
    
    def __append_tables_info(self, archive_file, inner_file_name, inner_file_ext):
        try:
            inner_file = archive_file.read(inner_file_name)
            binary = io.BytesIO(inner_file).read()
        except (BadZipfile, BadRarFile, NotRarFile, io.UnsupportedOperation):
            raise TypeError(f'{inner_file_ext}-архив не может быть прочитан')
        
        if inner_file_ext == 'zip':
            file = archive_file.read(inner_file_name)
            try:
                file_read = io.BytesIO(file)
                zip_file = ZipFile(file_read, 'r')
                for file_name in zip_file.namelist():
                    file_ext = self.__get_file_ext(file_name)
                    self.__append_tables_info(zip_file, file_name, file_ext)
            except BadZipfile:
                raise TypeError('zip-архив не может быть прочитан')
            
        elif inner_file_ext == 'rar':
            file = archive_file.read(inner_file_name)
            try:
                file_read = io.BytesIO(file)
                rar_file = RarFile(file_read)
                for file_name in rar_file.namelist():
                    file_ext = self.__get_file_ext(file_name)
                    self.__append_tables_info(rar_file, file_name, file_ext)
            except (BadRarFile, NotRarFile, io.UnsupportedOperation):
                raise TypeError('rar-архив не может быть прочитан')
        
        parser = self.__choose_parser(inner_file_ext)(binary=binary, html=None)
        tables_info = parser.get_tables_info()
        for ti in tables_info:
            self.tables_info.append(ti)
    
    def __get_file_ext(self, full_file_name):
        full_file_name = full_file_name.lower()
        splitted = full_file_name.split('.')
        file_ext = None
        if len(splitted) > 1: # в противном случае наткнулись на название директории, а не файла
            file_ext = splitted[-1]
        return file_ext
    
    def __choose_parser(self, file_ext):
        parsers_by_ext = {
            'xlsx': ParserXLSX,
            'xls': ParserXLS,
            'docx': ParserDOCX,
            'zip': ParserArchive,
            'rar': ParserArchive,
            'htm': ParserHTM,
        }
        
        if file_ext in parsers_by_ext:
            return parsers_by_ext[file_ext]
        else:
            raise TypeError(f'подходящий парсер не найден (расширение {file_ext})')


class TableObject:
    graph_id = None
    path = None
    name = None
    n_rows = None
    n_columns = None
    unit = None
    number = None

    def __init__(self, idx):
        self.idx = idx


file_types = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ParserXLSX,
    'application/vnd.ms-excel': ParserXLS,
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': ParserDOCX,
    'application/zip': ParserArchive,
    'application/x-zip-compressed': ParserArchive,
    'application/x-rar-compressed': ParserArchive,
    'text/html; charset=UTF-8': ParserHTM,
    'text/html; charset=utf-8': ParserHTM,
    'text/html; charset="utf-8"': ParserHTM,
    'text/html; charset=windows-1251': ParserHTM,
    'text/html; charset=cp1251': ParserHTM,
    'text/html; charset=koi8-r': ParserHTM,
    'text/html': ParserHTM,
    'htm': ParserHTM,
}


def choose_parser(graph_node):
    obj_type = graph_node.type
    obj_path = graph_node.path
    obj_ext = obj_path.split('.')[-1].lower()
    obj_binary = graph_node.file
    obj_html = graph_node.document
    
    if obj_type in GRAPH_ERRORS:
        raise TypeError(f'узел графа не собран ({obj_type})')
    elif obj_type in file_types:
        return file_types[obj_type](obj_binary, obj_html)
    elif obj_ext in file_types:
        return file_types[obj_ext](obj_binary, obj_html)
    else:
        raise TypeError(f'подходящий парсер не найден (тип объекта {obj_type}, расширение {obj_ext})')
